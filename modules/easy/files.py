import pandas as pd
import openpyxl
import os


class Excel:
    def make_excel(user_filename, server_filename, rows, cols):
        filepath = os.path.join('uploads', server_filename)
        dataframe = pd.DataFrame(rows, columns=cols)
        writer = pd.ExcelWriter(filepath, engine="xlsxwriter")
        dataframe.to_excel(writer, sheet_name=user_filename,index=False)
        writer.close()
        wb = openpyxl.load_workbook(filepath)
        ws = wb[user_filename]
        max_col = ws.max_column
        for col in range(1, max_col+1):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = 25
        wb.save(filepath)
        return filepath

    def delete_file(filepath):
        os.remove(filepath)